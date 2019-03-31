# -*- coding: utf-8 -*-
"""
Created on Tue Mar  5 12:11:03 2019

@author: hxh
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Mar  4 21:12:56 2019

@author: hxh
"""

import random
import openpyxl
from datetime import datetime
import re

def init(s):#初始化矩阵，根据不同的相似度有不同的矩阵,s是相似度
    temp=[[] for i in range (15)]
    temp1=[]
    for i in range (15):#初始化最初的15*15的矩阵
        for j in range (15):
            if (abs(i-j)<5):
                temp1.append(i+1)
                temp1.append(j+1)
            else:
                temp1.append(0)
            temp[i].append(temp1)
            temp1=[]
   
    if (s==1):#判断s的值
        x=4#x是删除次数
        while (x!=0):
            r=random.randint(3,14)#行数为4到15
            y=random.randint(0,11)#列数为1到12
            while (r>y and abs(r-y)>2):
                r=random.randint(3,14)
                y=random.randint(0,11)
            if (temp[r][y]!=[0]):
                x-=1
                temp[r][y]=[0]
                print ("这次删除的是",r+1,"行",y+1,"列的元素")
        return temp
    elif (s==2):
        x=3
        while (x!=0):
            r=random.randint(0,14)
            y=random.randint(0,14)
            while ((r-y)>0):
                r=random.randint(0,14)
                y=random.randint(0,14)
            if (temp[r][y]!=[0]):
                x-=1
                temp[r][y]=[0]
                print ("这次删除的是",r+1,"行",y+1,"列的元素")
        return temp
    elif (s==3):
        x=2
        while (x!=0):
            r=random.randint(0,14)
            y=random.randint(0,14)
            while ((r-y)>0):
                r=random.randint(0,14)
                y=random.randint(0,14)
            if (temp[r][y]!=[0]):
                x-=1
                temp[r][y]=[0]
                print ("这次删除的是",r+1,"行",y+1,"列的元素")
        return temp
    elif (s==4):
        x=3
        while (x!=0):
            r=random.randint(0,14)
            y=random.randint(0,14)
            while ((r-y)>=0):
                r=random.randint(0,14)
                y=random.randint(0,14)
            if (temp[r][y]!=[0]):
                x-=1
                temp[r][y]=[0]
                print ("这次删除的是",r+1,"行",y+1,"列的元素")
        return temp
    else:
        return temp
   

def ran(temp,a):#从矩阵中随机选一个元素，返回行和列的差,a为行为，即行数
    r=a
    y=random.randint(0,14)
    while (temp[r][y]==[0]):#如果这个位置的元素是[0]，说明该位置元素已经被删除了
        y=random.randint(0,14)
    return (r-y)

def put(row,col,act):#把概率存放到excel中,row是行，col是列,act是行为
    wb=openpyxl.load_workbook("D:/xianyupro/juzhen.xlsx")
    sheetnames=wb.get_sheet_names()
    ws=wb.get_sheet_by_name(sheetnames[0])
    ws.cell(row,col).value=act
    wb.save("D:/xianyupro/juzhen.xlsx")
    
def ave(count):#求平均数
    c=0
    wb=openpyxl.load_workbook("D:/xianyupro/juzhen.xlsx")
    sheetnames=wb.get_sheet_names()
    ws=wb.get_sheet_by_name(sheetnames[0])
    sum=0
    for i in ws.rows:
        for j in i:
            sum+=j.value
        ws.cell(11+c,1).value=sum/(count-1)
        c+=1
        sum=0
    wb.save("D:/xianyupro/juzhen.xlsx")
       
def cle():#每次程序启动前，清除excel表
    wb = openpyxl.Workbook()
    sheetnames=wb.get_sheet_names()
    ws=wb.get_sheet_by_name(sheetnames[0])
    wb.remove_sheet(ws)
    wb.create_sheet('Sheet1')    
    wb.save("D:/xianyupro/juzhen.xlsx")


if __name__=='__main__':
    cle()
    s=input('请输入相似度：1-5\n')
    s=int(s)
    temp=init(s)#s是相似度,temp为删减过的矩阵
    print("请输入行为1，概率1，行为2，概率2,行为用1-15概率用小数表示，比如0.75\n")
    x1=int(input())-1
    x2=float(input())
    x3=int(input())-1
    x4=float(input())
    time=int(input('请输入次数\n'))
    begintime=datetime.now()
    time1=int(time*x2)#行为1的次数
    time2=int(time*x4)#行为2的次数
    a1,a2,a3,a4,a5,a6,a7,a8,a9=0,0,0,0,0,0,0,0,0
    b1,b2,b3,b4,b5,b6,b7,b8,b9=0,0,0,0,0,0,0,0,0
    count=1
    
    sum=0#用于求和    
    tt=[]#用于存放随机选取的操作

    for i in range(time1):
        sub=ran(temp,x1)
        print("此次操作是",x1+1)
        sum+=x1+1
        tt.append(x1+1)
        if (sub==4):
            a1+=1
        elif(sub==3):
            a2+=1
        elif(sub==2):
            a3+=1
        elif(sub==1):
            a4+=1
        elif(sub==0):
            a5+=1
        elif(sub==-1):
            a6+=1
        elif(sub==-2):
            a7+=1
        elif(sub==-3):
            a8+=1
        elif(sub==-4):
            a9+=1
        b1=a1/count
        b2=a2/count
        b3=a3/count
        b4=a4/count
        b5=a5/count
        b6=a6/count
        b7=a7/count
        b8=a8/count
        b9=a9/count
        print("从1-9的斜对角的概率是：",b1,b2,b3,b4,b5,b6,b7,b8,b9)
        print("\n\n")
        
        put(1,count,b1)
        put(2,count,b2)
        put(3,count,b3)
        put(4,count,b4)
        put(5,count,b5)
        put(6,count,b6)
        put(7,count,b7)
        put(8,count,b8)
        put(9,count,b9)
        
        count+=1
        
    for i in range (time2):
        sub=ran(temp,x3)
        print("此次操作是",x3+1)
        sum+=x3+1
        tt.append(x3+1)
        if (sub==4):
            a1+=1
        elif(sub==3):
            a2+=1
        elif(sub==2):
            a3+=1
        elif(sub==1):
            a4+=1
        elif(sub==0):
            a5+=1
        elif(sub==-1):
            a6+=1
        elif(sub==-2):
            a7+=1
        elif(sub==-3):
            a8+=1
        elif(sub==-4):
            a9+=1
        b1=a1/count
        b2=a2/count
        b3=a3/count
        b4=a4/count
        b5=a5/count
        b6=a6/count
        b7=a7/count
        b8=a8/count
        b9=a9/count
        print("从1-9的斜对角的概率是：",b1,b2,b3,b4,b5,b6,b7,b8,b9)
        print("\n\n")
        
        put(1,count,b1)
        put(2,count,b2)
        put(3,count,b3)
        put(4,count,b4)
        put(5,count,b5)
        put(6,count,b6)
        put(7,count,b7)
        put(8,count,b8)
        put(9,count,b9)
        
        count+=1
    for i in range(time-time1-time2):
        xx=random.randint(0,14)#剩下的行为为15次除了前两次的
        while(xx==x1 or xx==x2):
            xx=random.randint(0,14)
        print("这次随机选取到的操作是",xx+1)
        sum+=xx+1
        tt.append(xx+1)
        sub=ran(temp,xx)
        if (sub==4):
            a1+=1
        elif(sub==3):
            a2+=1
        elif(sub==2):
            a3+=1
        elif(sub==1):
            a4+=1
        elif(sub==0):
            a5+=1
        elif(sub==-1):
            a6+=1
        elif(sub==-2):
            a7+=1
        elif(sub==-3):
            a8+=1
        elif(sub==-4):
            a9+=1
        b1=a1/count
        b2=a2/count
        b3=a3/count
        b4=a4/count
        b5=a5/count
        b6=a6/count
        b7=a7/count
        b8=a8/count
        b9=a9/count
        print("从1-9的斜对角的概率是：",b1,b2,b3,b4,b5,b6,b7,b8,b9)
        print("\n\n")
        
        put(1,count,b1)
        put(2,count,b2)
        put(3,count,b3)
        put(4,count,b4)
        put(5,count,b5)
        put(6,count,b6)
        put(7,count,b7)
        put(8,count,b8)
        put(9,count,b9)
        
        count+=1
    print("随机选取的行为次数为",(time-time1-time2),"\n全部的操作分别是:",tt,"\n全部的和为：\n",sum)
    ave(count)
    finishtime=datetime.now()
    totaltime=finishtime-begintime
    totaltime=str(totaltime)
    ttt=re.findall(r'\d+',totaltime)
    hours=ttt[0]
    minutes=ttt[1]
    seconds=re.findall(r'\d+\.\d+',totaltime)
    print("此次程序共用时：",hours,"时",minutes,"分",seconds[0],"秒")