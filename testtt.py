# -*- coding: utf-8 -*-
"""
Created on Sat Mar  2 21:31:53 2019

@author: hxh
"""
import openpyxl
import xlrd
import re
import xlrd
from xlutils.copy import copy

def test1(x):
    wb=openpyxl.load_workbook("D:/xianyupro/author1.xlsx")
    sheetnames=wb.get_sheet_names()
    ws=wb.get_sheet_by_name(sheetnames[0])
    return (ws.cell(1,x).value)

def test2(x,n):
    wb=openpyxl.load_workbook("D:/xianyupro/author1.xlsx")
    sheetnames=wb.get_sheet_names()
    ws=wb.get_sheet_by_name(sheetnames[0])
    ws.cell(1,x).value=n
    wb.save("D:/xianyupro/author1.xlsx")

def test3():
    wb=openpyxl.load_workbook("D:/xianyupro/author1.xlsx")
    sheetnames=wb.get_sheet_names()
    ws=wb.get_sheet_by_name(sheetnames[0])
    for i in range (4):
        ws.cell(1,i+1).value=0
    wb.save("D:/xianyupro/author1.xlsx")

def upgrade(aut,name):
    wb=xlrd.open_workbook("D:/xianyupro/author.xls")    
    ws=wb.sheet_by_name(name)
    temp=[]
    temp1=[]
    aut1=aut
    for i in range(2,6):
        for j in range(ws.nrows-5,ws.nrows):
            temp.append(ws.cell(j,i).value)
    for i in range(len(aut)):
        for j in range (5*i,5*(i+1)):
            temp1.append(temp[j])
        if(type(max(temp1))==float):
            if (max(temp1)<0.5):#5次中使用了一次或更少就删去该功能
                aut1[i]=0
        temp1=[]
    while 0 in aut1:
        aut1.remove(0)
    return (aut1)#返回一个新的权限

def repeat(name):#判断该表是不是重复
    wb=xlrd.open_workbook("D:/xianyupro/author.xls")
    newb=copy(wb)
    sheetnames=wb.sheet_names()#获取已经在excel中的全部表名，用于比较是否插入新表
    for i in range (len(sheetnames)):
        if (name==sheetnames[i]):
            newb.save("D:/xianyupro/author.xls")
            return 0
    newb.save("D:/xianyupro/author.xls")
    return 1

'''
if __name__=='__main__':
    name=input('请输入用户名')
    if (repeat(name)==1):
        print("不存在")
    else:
        print("该用户已存在，请输入其他用户名")
'''